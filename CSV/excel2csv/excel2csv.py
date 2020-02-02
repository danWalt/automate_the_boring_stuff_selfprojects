#! python3

# excel2csv.py filters a directory for excel files and saves each sheet in
# an excel file as a separate CSV file

import openpyxl, csv, os

os.makedirs('csvFiles', exist_ok=True)

for excelFile in os.listdir('.'):
    excel_file_name = excelFile[:len(excelFile) - 5]
    # Skip non-xlsx files, load the workbook object.
    if excelFile.endswith('.xlsx'):
        wb = openpyxl.load_workbook(excelFile)
        # Loop through every sheet in the workbook.
        for sheetName in wb.sheetnames:
            sheet = wb[sheetName]

            # Create the CSV filename from the Excel filename and sheet title.
            csvfilename = excel_file_name + '_' + sheetName + '.csv'
            csvFileObj = open(os.path.join('csvFiles', csvfilename), 'w',
                              newline='')
            # Create the csv.writer object for this CSV file.
            csvWriter = csv.writer(csvFileObj)

            # Loop through every row in the sheet.
            for rowNum in range(1, sheet.max_row + 1):
                rowData = []  # append each cell to this list
                # Loop through each cell in the row.
                for columnNum in range(1, sheet.max_column + 1):
                    # Append each cell's data to rowData.
                    rowData.append(
                        sheet.cell(row=rowNum, column=columnNum).value)
                # Write the rowData list to the CSV file.
                csvWriter.writerow(rowData)
            csvFileObj.close()
