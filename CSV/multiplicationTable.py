#! python3
import sys, openpyxl
from openpyxl.styles import Font

# make sure that an argument was given
if len(sys.argv) != 2:
    print('Exactly 1 argument is to be passed to the script')
# create a new workbook
wb = openpyxl.Workbook()
# load the first worksheet in the new workbook
sheet = wb.active
# feel in the 1st column and row
for i in range(2, int(sys.argv[1]) + 2):
    currentCell = sheet.cell(row=1, column=i)
    currentCell.value = i - 1
    currentCell.font = Font(bold=True)
    currentCell = sheet.cell(row=i, column=1)
    currentCell.value = i - 1
    currentCell.font = Font(bold=True)

# multiply the values
for i in range(2, int(sys.argv[1]) + 2):
    for j in range(2, int(sys.argv[1]) + 2):
        currentCell = sheet.cell(row=i, column=j)
        currentCell.value = (i - 1) * (j - 1)

# save the workbook in a new xlsx file
wb.save('multiplicationTabel.xlsx')
